import json
import os
import re
import random
import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill

import ollama
from logging_utils import log_and_print
from config import (
    OLLAMA_MODEL,
    OLLAMA_OPTIONS,
    OLLAMA_PROMPT_MODE,
    OFFLINE_QUEUE_FILE,
    EXCEL_FILE,
    FALLBACK_TEXT_FILE,
    EXCEL_COLOR_LIST)
last_chosen_color = None

class OllamaAIChat:
    """
    AI component for sending raw text blocks to Ollama using the generate API
    and logging actionable tasks into an Excel sheet.
    """
    def __init__(self, model=OLLAMA_MODEL, options=None):
        if options is None:
            options = OLLAMA_OPTIONS
        self.model = model
        self.options = options
        try:
            self.client = ollama.Client()
            log_and_print(f"OllamaAIChat initialized with model '{self.model}'.")
        except Exception as e:
            log_and_print(f"Failed to initialize Ollama client: {e}")
            raise e

    def _build_prompt(self, raw_text):
        """
        Builds a prompt for Ollama based on the mode specified in the config.
        
        - If OLLAMA_PROMPT_MODE is set to "restrictive", instructs Ollama to extract only the most critical details
        and return a concise JSON array of bullet points.
        - Otherwise, it uses the default summarization prompt to generate a comprehensive bullet-point summary.
        """
        if OLLAMA_PROMPT_MODE.lower() == "restrictive":
            prompt = (
                "Your job is to analyze a block of text provided to you at the end of this prompt, extract the explicit details, and organize those details into bullet points.\n"
                "Each bullet point should use the full context of the text to ensure clarity. For example, replace vague references like 'it' or 'they' with the actual subject or object they refer to.\n"
                "Your job is NOT to not summarize with creative freedom.\n"
                "Your job is NOT to break the provided block of text into individual words.\n"
                "Ensure every single detail in the provided block of text is captured and converted into a bullet point.\n\n"
                "Your output should:\n"
                "- Be organized into a JSON array, where each bullet point is a separate string. Example: [\"bullet point 1\", \"bullet point 2\", \"bullet point 3\"]\n"
                "- Contain nothing but the JSON array of bullet points.\n"
                "- Exclude any additional commentary, explanations, or formatting outside the JSON array.\n"
                "- Avoid backticks (`), code block formatting (e.g., ```json), or extra symbols.\n"
                "- Only return a plain JSON array of strings, nothing else.\n\n"
                f"Here is the text to analyze:\n\"\"\"{raw_text}\"\"\"\n\n"
                "Now return a JSON array of bullet points:"
            )
        else:
            prompt = (
                "Your job is to analyze a block of text provided to you at the end of this prompt and transform it into a comprehensive bullet-point summary.\n"
                "Your job is NOT to break the provided block of text into individual words.\n"
                "Ensure every single detail in the provided block of text is captured and converted into a bullet point.\n\n"
                "Each bullet point should:\n"
                "- Be clear and concise, summarizing the text effectively while retaining all relevant details.\n"
                "- Use the full context of the text to ensure clarity. For example, replace vague references like 'it' or 'they' with the actual subject or object they refer to.\n"
                "- NOT be a duplicate of the previous bullet point\n\n"
                "Your output should:\n"
                "- Be organized into a JSON array, where each bullet point is a separate string. Example: [\"bullet point 1\", \"bullet point 2\", \"bullet point 3\"]\n"
                "- Contain nothing but the JSON array of summarized bullet points.\n"
                "- Exclude any additional commentary, explanations, or formatting outside the JSON array.\n"
                "- Avoid backticks (`), code block formatting (e.g., ```json), or extra symbols.\n"
                "- Only return a plain JSON array of strings, nothing else.\n\n"
                f"Here is the text to analyze:\n\"\"\"{raw_text}\"\"\"\n\n"
                "Now return a JSON array of bullet points:"
            )
        return prompt

    def _build_fix_prompt(self, raw_response):
        """
        Prompt used if bracket extraction or JSON parse fails.
        Instructs Ollama to fix the original text to become valid JSON.
        """
        prompt = (
            "Your job is to fix the provided text, which was supposed to be a JSON array but is not valid. Your job is to turn the provided text into a valid JSON array\n"
            "You must remove, add, or modify any characters so the final output is:\n"
            "- Only a valid JSON array of strings.\n"
            "- Nothing else; no code fences, no extra commentary.\n\n"
            "Here is the invalid text:\n\n"
            f"\"\"\"{raw_response}\"\"\"\n\n"
            "Fix it so it becomes a valid JSON array of strings, and return only the fixed JSON array."
        )
        return prompt

    def _attempt_fix_prompt(self, raw_response):
        """
        Single helper function that re-feeds 'raw_response' to Ollama
        using the fix prompt, then attempts bracket extraction again.
        If bracketed_fix is "[]", return None to signal a dead-end.
        """
        fix_prompt = self._build_fix_prompt(raw_response)
        try:
            fix_data = self.client.generate(prompt=fix_prompt, model=self.model, options=self.options)
            log_and_print(f"\n[OllamaAIChat] Second Attempt Fix Prompt Response:\n{fix_data}")
            fixed_response = fix_data.get("response", "")

            # Attempt bracket extraction again
            match_fix = re.search(r"\[.*\]", fixed_response, flags=re.DOTALL)
            if match_fix:
                bracketed_fix = match_fix.group(0)
            else:
                if '[' in fixed_response and ']' not in fixed_response:
                    attempt_fix = fixed_response + "]"
                    match_fix2 = re.search(r"\[.*\]", attempt_fix, flags=re.DOTALL)
                    bracketed_fix = match_fix2.group(0) if match_fix2 else "[]"
                elif ']' in fixed_response and '[' not in fixed_response:
                    attempt_fix = "[" + fixed_response
                    match_fix2 = re.search(r"\[.*\]", attempt_fix, flags=re.DOTALL)
                    bracketed_fix = match_fix2.group(0) if match_fix2 else "[]"
                else:
                    bracketed_fix = "[]"

            bracketed_fix = " ".join(bracketed_fix.splitlines())
            bracketed_fix = re.sub(r'[\n\r\t]', '', bracketed_fix)

            # If bracketed_fix is still "[]", we treat this as final failure.
            if bracketed_fix.strip() == "[]":
                log_and_print("[OllamaAIChat] Fix prompt returned an empty array. Stopping here.")
                return None  # indicates we should stop further checks

            return bracketed_fix
        except Exception as e2:
            log_and_print(f"[OllamaAIChat] Error during second attempt fix call: {e2}")
            return None

    def _pulse_check_ollama(self):
        """
        Sends a minimal prompt to check if the Ollama client is online.
        Returns True if the client responds without error, False otherwise.
        """
        try:
            test_prompt = "ping"
            # Sending a minimal prompt; we don't need a complex response.
            self.client.generate(prompt=test_prompt, model=self.model, options=self.options)
            return True
        except Exception as e:
            log_and_print(f"[OllamaAIChat] Pulse check failed: {e}")
            return False

    
    def _store_offline_block(self, block_id, raw_text):
        """
        Appends a single block to the offline queue file as a JSON line:
          {"block_id": 123, "raw_text": "..."}
        """
        obj = {"block_id": block_id, "raw_text": raw_text}
        line = json.dumps(obj, ensure_ascii=False)
        with open(OFFLINE_QUEUE_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")

        log_and_print(f"[OllamaAIChat] Stored block {block_id} offline in {OFFLINE_QUEUE_FILE}.")
    
    def _try_offline_queue(self):
        """
        Attempts to load and process queued blocks from OFFLINE_QUEUE_FILE in order.
        Instead of stopping at the first failure, it processes every line.
        Successfully processed blocks are removed from the file, while any blocks
        that still fail are written back to the file.
        """
        if (not os.path.exists(OFFLINE_QUEUE_FILE)) or (os.path.getsize(OFFLINE_QUEUE_FILE) == 0):
            log_and_print("Offline queue is empty or offline queue file does not exist yet")
            return  # No offline queue yet or it's empty
        
        # Pulse check to verify if Ollama is available
        if not self._pulse_check_ollama():
            log_and_print("[OllamaAIChat] Pulse check failed. Skipping offline queue processing.")
            return
        else:
            log_and_print("[OllamaAIChat] Pulse check passed. Processing offline queue...")

        # Read all lines once
        with open(OFFLINE_QUEUE_FILE, "r", encoding="utf-8") as f:
            lines = f.readlines()

        still_failing = []
        total_count = 0
        success_count = 0

        for line in lines:
            line = line.strip()
            if not line:
                continue
            total_count += 1
            try:
                obj = json.loads(line)
                block_id = obj["block_id"]
                raw_text = obj["raw_text"]
            except (json.JSONDecodeError, KeyError):
                # If the line is corrupted, skip it
                continue

            log_and_print(f"[OllamaAIChat] Re-processing offline block {block_id} from queue. Raw transcribed text: {raw_text}")
            try:
                self._process_block_internal(raw_text, block_id)
                success_count += 1
            except Exception as e:
                log_and_print(f"[OllamaAIChat] Block {block_id} still failing: {e}.")
                still_failing.append(line)

        # Re-write the offline queue file with all still failing blocks
        with open(OFFLINE_QUEUE_FILE, "w", encoding="utf-8") as f:
            for fail_line in still_failing:
                f.write(fail_line + "\n")

        if total_count > 0:
            log_and_print(f"[OllamaAIChat] Retried {total_count} queued blocks. {success_count} succeeded, {total_count - success_count} remain failing.")
    
    def process_block(self, raw_text, block_id):
        """
        Public method. On each new block from the transcription worker:
        1) Process any offline queue blocks.
        2) Then process this new block. If that fails, store it offline.
        """
        # 1) First, process any offline queue blocks
        self._try_offline_queue()

        # 2) Now process THIS block
        try:
            self._process_block_internal(raw_text, block_id)
        except Exception as e:
            log_and_print(f"[OllamaAIChat] Block {block_id} failed: {e}. Storing offline.")
            self._store_offline_block(block_id, raw_text)

    def _process_block_internal(self, raw_text, block_id):
        """
        The actual logic that calls Ollama once for a given block
        (extracted from your original 'process_block' method).
        If it fails, it raises an exception, which our caller can handle
        by storing offline or ignoring.
        """
        log_and_print(f"\n[OllamaAIChat] Processing block ID={block_id}, length={len(raw_text)} chars.")

        prompt = self._build_prompt(raw_text)
        try:
            response_data = self.client.generate(prompt=prompt, model=self.model, options=self.options)

            #printing and logging filtered resonse without "context" object as it's lengthy and unused
            try:
                # Convert response_data to a dictionary
                response_dict = vars(response_data)
                # Create a new dictionary without the "context" key
                filtered_response_data = {k: v for k, v in response_dict.items() if k != "context"}
                # Print the filtered response
                log_and_print(f"\n[OllamaAIChat] Full Response from Ollama:\n{filtered_response_data}")

            except Exception as e:
                # If an error occurs, print the full response without filtering
                log_and_print(f"\n[OllamaAIChat] Failed to filter response. Printing full response instead:\n{response_data}")
                log_and_print(f"Error: {e}")

            raw_response = response_data.get("response", "")
            # The same bracket extraction + fix prompt logic as before...
            match = re.search(r"\[.*\]", raw_response, flags=re.DOTALL)
            if match:
                bracketed_content = match.group(0)
            else:
                if '[' in raw_response and ']' not in raw_response:
                    attempt = raw_response + "]"
                    match2 = re.search(r"\[.*\]", attempt, flags=re.DOTALL)
                    bracketed_content = match2.group(0) if match2 else "[]"
                elif ']' in raw_response and '[' not in raw_response:
                    attempt = "[" + raw_response
                    match2 = re.search(r"\[.*\]", attempt, flags=re.DOTALL)
                    bracketed_content = match2.group(0) if match2 else "[]"
                else:
                    bracketed_content = self._attempt_fix_prompt(raw_response)
                    if bracketed_content is None:
                        raise Exception("Fix prompt returned None (dead-end).")

            # Clean up
            bracketed_content = " ".join(bracketed_content.splitlines())
            bracketed_content = re.sub(r'[\n\r\t]', '', bracketed_content)

            # Attempt to parse
            try:
                tasks = json.loads(bracketed_content.strip())
                if isinstance(tasks, list) and all(isinstance(item, str) for item in tasks):
                    log_and_print(f"[OllamaAIChat] Parsed Tasks:\n{tasks}")
                    self._log_tasks_to_excel(block_id, tasks)
                else:
                    log_and_print("[OllamaAIChat] The response is not a valid JSON list of strings. Attempting to fix...")
                    bracketed_fix = self._attempt_fix_prompt(raw_response)
                    if bracketed_fix is None:
                        raise Exception("Fix prompt returned None (structure fix).")

                    try:
                        tasks2 = json.loads(bracketed_fix.strip())
                        if isinstance(tasks2, list) and all(isinstance(x, str) for x in tasks2):
                            log_and_print(f"[OllamaAIChat] Parsed Fixed Tasks (structure fix):\n{tasks2}")
                            self._log_tasks_to_excel(block_id, tasks2)
                        else:
                            raise Exception("Even after fix prompt, not a valid JSON list of strings (structure).")
                    except json.JSONDecodeError:
                        raise Exception("Fix prompt failed to parse JSON again (structure).")

            except json.JSONDecodeError:
                log_and_print("[OllamaAIChat] Failed to parse Ollama's response as JSON. Attempting to fix...")
                bracketed_fix = self._attempt_fix_prompt(raw_response)
                if bracketed_fix is None:
                    raise Exception("Fix prompt returned None (parse).")

                try:
                    tasks2 = json.loads(bracketed_fix.strip())
                    if isinstance(tasks2, list) and all(isinstance(item, str) for item in tasks2):
                        log_and_print(f"[OllamaAIChat] Parsed Fixed Tasks:\n{tasks2}")
                        self._log_tasks_to_excel(block_id, tasks2)
                    else:
                        raise Exception("Even after fix prompt, not a valid JSON list of strings.")
                except json.JSONDecodeError:
                    raise Exception("Second attempt also failed to parse JSON.")

        except Exception as e:
            # Raise so that calling code can handle it (by storing offline).
            raise e

    def _log_tasks_to_excel(self, block_id, tasks):
        """
        Tasks have been renamed to Bullet point summary in the generated excel spreadsheet.
        Originally this porgram was focuesd only on extracting tasks/objectives from speech.
        Now it's programmed to summarizes everything into a bullet point style summary.
        """
        global last_chosen_color

        excel_file = EXCEL_FILE
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active
                log_and_print("[OllamaAIChat] Excel file loaded.")
            except FileNotFoundError:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(["Timestamp", "Block ID", "Bullet point summary"])
                log_and_print("[OllamaAIChat] Excel file created.")

            start_row = sheet.max_row + 1
            for task in tasks:
                sheet.append([now, block_id, task])

            end_row = sheet.max_row
            # Exclude last_chosen_color
            available_colors = [c for c in EXCEL_COLOR_LIST if c != last_chosen_color]
            chosen_color = random.choice(available_colors)
            last_chosen_color = chosen_color

            fill = PatternFill(start_color=chosen_color, end_color=chosen_color, fill_type="solid")
            for row_idx in range(start_row, end_row + 1):
                for col_idx in range(1, 4):
                    sheet.cell(row=row_idx, column=col_idx).fill = fill

            workbook.save(excel_file)
            log_and_print(f"[OllamaAIChat] Logged {len(tasks)} tasks to Excel file '{excel_file}' with color {chosen_color}.")

        except PermissionError:
            """
            This indicates the Excel file is open/locked. We can't save changes.
            We'll append tasks to a fallback text file that remains the same 
            throughout the program session, stored in FALLBACK_TEXT_FILE.
            """
            log_and_print(f"[OllamaAIChat] Excel file is locked. Falling back to '{FALLBACK_TEXT_FILE}'.")
            try:
                with open(FALLBACK_TEXT_FILE, "a", encoding="utf-8") as fallback:
                    for task in tasks:
                        fallback.write(f"{now}, {block_id}, {task}\n")
                log_and_print(f"[OllamaAIChat] Logged {len(tasks)} tasks to fallback file '{FALLBACK_TEXT_FILE}' until Excel is closed.")
            except Exception as e:
                log_and_print(f"[OllamaAIChat] Error logging tasks to fallback text file: {e}")

        except Exception as e:
            log_and_print(f"[OllamaAIChat] Error logging tasks to Excel: {e}")
